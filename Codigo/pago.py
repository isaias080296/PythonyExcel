import openpyxl

import subprocess
import sys
import pandas as pd
from pandas import ExcelWriter
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime, timedelta

print("\n")
sep = '|{}|'.format('-'*55)
print ('*********************************************************')
print('{0}\n|Hola Bienvenido al Sistema de Pago de Mansion Mex 2018 |\n{0}'.format(sep))
print ('*********************************************************')
print('{0}\n                 Que desea Realizar       \n{0}'.format(sep))
print("| 1 - Establecer Pago por Hora                          |")
print("| 2 - Imprimir Pago de Trabajadores                     |")
print("| 3 - Buscar Trabajador                                 |")
print("| 4 - Imprimir todos los trabajadores                   |")
print("| 5 - Guardar                                           |")
print("| 6 - Informacion del Programa                          |")
print("| 7 - Salir                                             |")
print("|_ _ _ _ _ _ _ _ _ _ _ _ _ _ _ _ _ _ _ _ _ _ _ _ _ _ _ _|")
print ("\n")

pago = float(input("Cual es el salario por Trabajador  "))
pago1= pago

doc1 = openpyxl.load_workbook("C:\Users\E5-575\Desktop\StandardReport.xlsx")
hoja = doc1.get_sheet_by_name('registro de pasar la tarjeta')


#
def restar_hora(hora1,hora2):
        formato = "%H:%M"
        h1 = datetime.strptime(hora1, formato)
        h2 = datetime.strptime(hora2, formato)
        resultado = h2 - h1
        return str(resultado)

def sumar_hora(hora1,hora2):
    formato = "%H:%M:%S"
    lista = hora2.split(":")
    hora=int(lista[0])
    minuto=int(lista[1])
    segundo=int(lista[2])
    h1 = datetime.strptime(hora1, formato)
    dh = timedelta(hours=hora)
    dm = timedelta(minutes=minuto)
    ds = timedelta(seconds=segundo)
    resultado1 =h1 + ds
    resultado2 = resultado1 + dm
    resultado = resultado2 + dh
    resultado=resultado.strftime(formato)
    return str(resultado)
def suma(hora1,hora2,hora3,minu1,minu2,minu3,pago):
        h1 = hora1
        h2 = hora2
        h3 = hora3
        m1=minu1
        m2=minu2
        m3=minu3
        pago1=((pago/7)/10)/60
        #print pago1
        resultado1 = ((h2+ h1+h3)*60 )+m1+m2+m3
        #print resultado1
        if(resultado1<=2400):
            resultado=(resultado1)
        elif(resultado1>2400):
            resultado=(resultado1+240)
            #print resultado
        return float(resultado)
#for columna in hoja.rows:
#     for fila in columna:
#                    print fila.value,
#     print ""
trabajadores =[]
dia1=[]
dia2=[]
dia3=[]
dia4=[]
dia5=[]
dia6=[]
dia7=[]
dia8=[]
dia1entrada=[]
dia2entrada=[]
dia3entrada=[]
dia4entrada=[]
dia5entrada=[]
dia6entrada=[]
dia7entrada=[]
dia8entrada=[]
dia1salida=[]
dia2salida=[]
dia3salida=[]
dia4salida=[]
dia5salida=[]
dia6salida=[]
dia7salida=[]
dia8salida=[]
dia1Total=[]
dia2Total=[]
dia3Total=[]
dia4Total=[]
dia5Total=[]
dia6Total=[]
dia7Total=[]
dia8Total=[]
diaP1=[]
diaP2=[]
diaP3=[]
diaP4=[]
STH=[]
diaM1=[]
diaM2=[]
diaM3=[]
diaM4=[]
STM=[]

diaTotal2=[]
diaTota34=[]
diaTota56=[]
diaTota78=[]
diaTota24=[]
diaTota68=[]
diaTota48=[]
seleccion = hoja['A4':'H4']
fechas =[]
for filas in seleccion:

#    sep = '|{}|{}|{}|{}|'.format('-'*6,  '-'*15,'-'*11,'-'*15)
#    print('{0}\n| Fecha| Horas por Dia | Trabajador| Total a Pagar |\n{0}'.format(sep))
    for columnas in filas:
                        dia=columnas.value
                        fechas +=[dia]
                        #print (' ', columnas.value)
    #print (fechas)
seleccion = hoja['K5':'K178']
i = 5
s=4
while i < 177:
 nombre=hoja.cell(row = i, column = 11).value
 trabajadores +=[nombre]
 if (i == 178):
   break
 i +=2
 seleccion = hoja['A5':'A178']
 ip1=6
 ip2=6
 ip3=6
 ip4=6
 ip5=6
 ip6=6
 ip7=6
 ip8 =6
 ip8entrada =6
while ip1 < 177:
  diauno=hoja.cell(row = ip1, column = 1).value
  dia1 +=[diauno]
  if (ip1 == 177):
    break
  ip1 +=2
while ip2 < 177:
  diados=hoja.cell(row = ip2, column = 2).value
  dia2 +=[diados]
  if (ip2 == 177):
    break
  ip2 +=2
while ip3 < 177:
  diatres=hoja.cell(row = ip3, column = 3).value
  dia3 +=[diatres]
  if (ip3 == 177):
    break
  ip3 +=2
while ip4 < 177:
  diacuatro=hoja.cell(row = ip4, column = 4).value
  dia4 +=[diacuatro]
  if (ip4 == 177):
    break
  ip4 +=2
while ip5 < 177:
  diacinco=hoja.cell(row = ip5, column = 5).value
  dia5 +=[diacinco]
  if (ip5 == 177):
    break
  ip5 +=2
while ip6 < 177:
  diaseis=hoja.cell(row = ip6, column = 6).value
  dia6 +=[diaseis]
  if (ip6 == 177):
    break
  ip6 +=2
while ip7 < 177:
  diasiete=hoja.cell(row = ip7, column = 7).value
  dia7 +=[diasiete]
  if (ip7 == 177):
    break
  ip7 +=2
while ip8 < 177:
  diaocho=hoja.cell(row = ip8, column = 8).value
  dia8 +=[diaocho]
  if (ip8 == 177):
    break
  ip8 +=2
null="0:00"
for ip8entrada in range(len(dia8)):
      diaocho1=str(dia8[ip8entrada])
      diao=diaocho1[0:5]
      diaocho1s=str(dia8[ip8entrada])
      diao1=diaocho1s[5:11]
      if diao=='None':
         dia8entrada+=[null]
         dia8salida +=[null]
      elif not diao1:
         dia8entrada+=[null]
         dia8salida +=[diao]
      else:
         dia8salida +=[diao]
         dia8entrada +=[diao1]
      ip8entrada +=1
#for ip8salida in range(len(dia8)):
#      diaocho1s=str(dia8[ip8salida])
#      diao1=diaocho1s
#      dia8salida +=[diao1]
#      ip8salida +=1
for ip1entrada in range(len(dia2)):
      diauno=str(dia1[ip1entrada])
      diao=diauno[0:5]
      diaocho1s=str(dia1[ip1entrada])
      diao1=diaocho1s[5:11]
      if diao=='None':
           dia1entrada+=[null]
           dia1salida +=[null]
      elif not (diao1) or not (diao):
           dia1entrada+=[null]
           dia1salida +=[null]

      else:
         dia1salida +=[diao1]
         dia1entrada +=[diao]
      ip1entrada +=1
for ip2entrada in range(len(dia2)):
      diauno=str(dia2[ip2entrada])
      diao=diauno[0:5]
      diaocho1s=str(dia2[ip2entrada])
      diao1=diaocho1s[5:11]
      if diao=='None':
           dia2entrada+=[null]
           dia2salida +=[null]
      elif not (diao1) or not (diao):
           dia2entrada+=[null]
           dia2salida +=[null]

      else:
         dia2salida +=[diao1]
         dia2entrada +=[diao]
      ip2entrada +=1
for ip3entrada in range(len(dia3)):
      diatres=str(dia3[ip3entrada])
      diao=diatres[0:5]
      diaocho1s=str(dia3[ip3entrada])
      diao1=diaocho1s[5:11]
      if diao=='None':
           dia3entrada+=[null]
           dia3salida +=[null]
      elif not (diao1) or not (diao):
           dia3entrada+=[null]
           dia3salida +=[null]
      else:
          dia3salida +=[diao1]
          dia3entrada +=[diao]
      ip3entrada +=1
for ip4entrada in range(len(dia4)):
      diatres=str(dia4[ip4entrada])
      diao=diatres[0:5]
      diaocho1s=str(dia4[ip4entrada])
      diao1=diaocho1s[5:11]
      if diao=='None':
           dia4entrada+=[null]
           dia4salida +=[null]
      elif not (diao1) or not (diao):
           dia4entrada+=[null]
           dia4salida +=[null]
      else:
          dia4salida +=[diao1]
          dia4entrada +=[diao]
      ip4entrada +=1
for ip5entrada in range(len(dia5)):
      diacinco=str(dia5[ip5entrada])
      diao=diacinco[0:5]
      diaocho1s=str(dia5[ip5entrada])
      diao1=diaocho1s[5:11]
      if diao=='None'or diao1 is None:
           dia5entrada+=[null]
           dia5salida +=[null]
      elif not (diao1) or not (diao):
           dia5entrada+=[null]
           dia5salida +=[null]

      else:
           dia5salida +=[diao1]
           dia5entrada +=[diao]
      ip5entrada +=1
for ip6entrada in range(len(dia6)):
      diaseis=str(dia6[ip6entrada])
      diao=diaseis[0:5]
      diaocho1s=str(dia6[ip6entrada])
      diao1=diaocho1s[5:11]
      if diao=='None':
           dia6entrada+=[null]
           dia6salida +=[null]
      elif not (diao1) or not (diao):
           dia6entrada+=[null]
           dia6salida +=[null]

      else:
          dia6salida +=[diao1]
          dia6entrada +=[diao]
      ip6entrada +=1
for ip7entrada in range(len(dia7)):
      diasiete=str(dia7[ip7entrada])
      diao=diasiete[0:5]
      diaocho1s=str(dia7[ip7entrada])
      diao1=diaocho1s[5:11]
      if diao=='None':
           dia7entrada+=[null]
           dia7salida +=[null]
      elif not (diao1) or not (diao):
           dia7entrada+=[null]
           dia7salida +=[null]

      else:
           dia7salida +=[diao1]
           dia7entrada +=[diao]
      ip7entrada +=1
#for ip7salida in range(len(dia7)):
#      diaocho1s=str(dia7[ip7salida])
#      diao1=diaocho1s[6:11]
#      dia7salida +=[diao1]
#      ip7salida +=1
#print dia1Total
#print dia1entrada
#print dia1salida
#print dia2entrada
#print dia2salida
#print dia3entrada
#print dia4entrada
#print dia5entrada
#print dia6entrada
#print dia7entrada
#print dia8entrada
#print dia1salida
#print dia3salida
#print dia4salida
#print dia5salida
#print dia6salida
#print dia7salida
#print dia8salida
#print dia1
#print dia2
#print dia3
#print dia4
#print dia5
#print dia6
#print dia7
#print dia8
opcion = input("Digite un numero ")
print ("\n")

subprocess.call("clear")
if opcion == 1:
         pago = input("Cual es la Hora de Pago por Trabajador")

elif opcion == 4:
      sep12 = '|{}|{}'.format('-'*7,'-'*18)
      print('{0}\n| Numero| Trabajador |\n{0}'.format(sep12))
      for i in range(len(trabajadores)):
                print("|  "+str(i)+ ".-    | "+str(trabajadores[i])+"|")
                print("")
                i+=1

elif opcion == 3:
        nomtrabajador = raw_input("Cual es el nombre del trabajador ")
        print (nomtrabajador in trabajadores)
elif opcion == 6:
          print ('**********************************************************************************************************************')
          print ('|Informacion del Sistema de Pago de Mansion Mex 2018                                                                 |')
          print ('**********************************************************************************************************************')
          print (' 1- El software toma los datos del Documento excel extraido del checador de huellas del Sistema (Steren)             |')
          print (' 2- Debe estar en la ruta Asignada  el documento con el nombre "StandardReport.xsl"                                  |')
          print (' 3- El documento estara vinculado con las Fechas de la Semana Creando un documento nuevo en excel para poder imprimir| ')
          print (' \n                                                                                                                     |')
          print (' Realizado por Isaias G. Gomez Duarte - Todos los derechos Reservados @MansionMex2018                                |')
elif opcion == 5:
# set file path
    il=1
    wb=Workbook()
    filepath="C:\Users\E5-575\Desktop\Pago de Trabajadores\Pago-Trabajadores Intamex"+str(fechas[0])+"-"+str(fechas[7])+".xlsx"
# save workbook
    wb.save(filepath)
    #seleccion1 = hoja['R2':'Y2']
    #print("Guardado con exito "+filepath)
# set file path

    wb=load_workbook(filepath)
# select demo.xlsx
    sheet=wb.active
# set value for cell A1=1seleccion = hoja['K5':'K157']
    s=2

    for il in range(len(trabajadores)):
               #print("  "+str(il)+ ".-     "+str(trabajadores[il]))
               #sheet.cell(row=s, column=1).value = str(trabajadores[il])
               #print("  "+str(il)+ ".-     "+str(dia1entrada[il]))
               #print(restar_hora(dia1entrada[il],dia1salida[il]))
               #sheet.cell(row=s, column=2).value =str(dia1entrada[il])
               #print("  "+str(il)+ ".-     "+str(dia1entrada[il]))
               #sheet.cell(row=s, column=3).value =str(dia1salida[il])
               #print("  "+str(il)+ ".-     "+str(dia1entrada[il]))
               #sheet.cell(row=s, column=4).value =str(dia2entrada[il])
               #print("  "+str(il)+ ".-     "+str(dia1entrada[il]))



               #sheet.cell(row=s, column=5).value =str(dia2salida[il])
               #sheet.cell(row=s, column=6).value =str(dia3entrada[il])
               #sheet.cell(row=s, column=7).value =str(dia3salida[il])
               #sheet.cell(row=s, column=8).value =str(dia4entrada[il])
               #sheet.cell(row=s, column=9).value =str(dia4salida[il])
               #sheet.cell(row=s, column=10).value =str(dia5entrada[il])
               #sheet.cell(row=s, column=11).value =str(dia5salida[il])
               #sheet.cell(row=s, column=12).value =str(dia6entrada[il])
               #sheet.cell(row=s, column=13).value =str(dia6salida[il])
               #sheet.cell(row=s, column=14).value =str(dia7entrada[il])
               #sheet.cell(row=s, column=15).value =str(dia7salida[il])
               #sheet.cell(row=s, column=16).value =str(dia8salida[il])
               #sheet.cell(row=s, column=17).value =str(dia8entrada[il])
               dia1Total +=[restar_hora(dia1entrada[il],dia1salida[il])]
               dia2Total +=[restar_hora(dia2entrada[il],dia2salida[il])]
               dia3Total +=[restar_hora(dia3entrada[il],dia3salida[il])]
               dia4Total +=[restar_hora(dia4entrada[il],dia4salida[il])]
               dia5Total +=[restar_hora(dia5entrada[il],dia5salida[il])]
               dia6Total +=[restar_hora(dia6entrada[il],dia6salida[il])]
               dia7Total +=[restar_hora(dia7entrada[il],dia7salida[il])]
               dia8Total +=[restar_hora(dia8entrada[il],dia8salida[il])]
               diaTotal2 +=[sumar_hora(dia1Total[il],dia2Total[il])]
               diaTota34 +=[sumar_hora(dia4Total[il],dia5Total[il])]
               diaTota56 +=[sumar_hora(dia6Total[il],dia7Total[il])]
               #diaTota78 +=[sumar_hora(dia8Total[il])]
               diaPA=str(diaTotal2[il])
               diaOO=diaPA[0:2]
               diaP1 +=map(int,[diaOO])
               diaPA1=str(diaTota34[il])
               diaOO1=diaPA1[0:2]
               diaP2 +=map(int,[diaOO1])
               diaPA2=str(diaTota56[il])
               diaOO2=diaPA2[0:2]
               diaP3 +=map(int,[diaOO2])
               diaA=str(diaTotal2[il])
               diaO=diaPA[3:5]
               diaM1 +=map(int,[diaO])
               diaA1=str(diaTota34[il])
               diaO1=diaA1[3:5]
               diaM2 +=map(int,[diaO1])
               diaA2=str(diaTota56[il])
               diaO2=diaA2[3:5]
               diaM3 +=map(int,[diaO2])
               if suma(diaP1[il],diaP2[il],diaP3[il],diaM1[il],diaM2[il],diaM3[il],pago)!=0:
                    sheet.cell(row=s, column=1).value = str(trabajadores[il])
                    sheet.cell(row=s, column=2).value =restar_hora(dia1entrada[il],dia1salida[il])
                    sheet.cell(row=s, column=3).value =restar_hora(dia2entrada[il],dia2salida[il])
                    sheet.cell(row=s, column=4).value =restar_hora(dia3entrada[il],dia3salida[il])
                    sheet.cell(row=s, column=5).value =restar_hora(dia4entrada[il],dia4salida[il])
                    sheet.cell(row=s, column=6).value =restar_hora(dia5entrada[il],dia5salida[il])
                    sheet.cell(row=s, column=7).value =restar_hora(dia6entrada[il],dia6salida[il])
                    sheet.cell(row=s, column=8).value =restar_hora(dia7entrada[il],dia7salida[il])
               #sheet.cell(row=s, column=8).value =restar_hora(dia8salida[il],dia8entrada[il])
               #sheet.cell(row=s, column=2).value =sumar_hora(dia1Total[il],dia2Total[il])
               #sheet.cell(row=s, column=3).value =sumar_hora(dia4Total[il],dia5Total[il])
               #sheet.cell(row=s, column=4).value =sumar_hora(dia6Total[il],dia7Total[il])
                    sheet.cell(row=s, column=9).value =suma(diaP1[il],diaP2[il],diaP3[il],diaM1[il],diaM2[il],diaM3[il],pago)
                    s+=1

               ###Total de Horas por Dia

               #diaTota24 +=[sumar_horas(diaTotal2[il],diaTota34[il])]
               #diaTota68 +=[sumar_horas(diaTota56[il],diaTota78[il])]
               #diaTota48 +=[sumar_hora(diaTota24[il],diaTota68[il])]

               #print("")

               il+=1

    sheet['A1'] = "Trabajadores"
# set value for cell B2=2
    sheet.cell(row=1, column=2).value = str(fechas[0])
    sheet.cell(row=1, column=3).value = str(fechas[1])
    sheet.cell(row=1, column=4).value = str(fechas[2])
    sheet.cell(row=1, column=5).value = str(fechas[3])
    sheet.cell(row=1, column=6).value = str(fechas[4])
    sheet.cell(row=1, column=7).value = str(fechas[5])
    sheet.cell(row=1, column=8).value = str(fechas[6])
    sheet.cell(row=1, column=9).value = "Pago Total"


# save workbook
    wb.save(filepath)
#print diaTotal2
#print diaTota34
#print diaTota56
#print diaTota78
#print diaTota24
